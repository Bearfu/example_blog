import os
import time
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

__author__ = 'bear_fu'


# 函数返回指定文件夹下所有的文件地址。
def all_path(dirname):
    result = []
    for maindir, subdir, file_name_list in os.walk(dirname):
        for filename in file_name_list:
            apath = os.path.join(maindir, filename)
            result.append(apath)
    return result


# 函数返回指定的路径下文件中的sheet名称列表
def fun_sheet_list(mk_path):
    print("正在检查Sheet名称是否存在问题")
    sheet_num = 0
    sheet_list = []
    for path in all_path(mk_path):
        # 若sheet数量为0（首次运行）
        if sheet_num is 0:
            # 将获取到的本文件的sheet数量赋值给sheet_num
            sheet_num = sheet(path)[0]
            # 若sheet的值与本文件的sheet数量不同
        elif sheet_num is not sheet(path)[0]:
            #  抛出文件的路径，结束循环
            print(path)
            return path
        else:  # sheet 的值与本文件的sheet数量相同
            sheet_list = sheet(path)[1:]
    return sheet_list


# 函数返回指定的路径下文件中的sheet名称列表
def sheet(excel_path):
    sheet_list = []
    wb = load_workbook(excel_path)
    count = len(wb.sheetnames)  # sheet数量
    sheet_list.append(count)
    for sheet in wb:
        sheet_list.append(sheet.title)  # sheet名称
    return sheet_list


# 函数返回指定的路径下文件中的指定字符串,指定的所在的坐标和组成的list
# 样式为 [["x","y","path"],["x","y","path"],["x","y","path"]]
def fun_page_position(excel_path, sheet_name):
    page_position_list = []
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name(sheet_name)
    nrows = ws.max_row  # 行数
    ncols = ws.max_column  # 列数
    for x_num in range(1, ncols + 1):
        col = get_column_letter(x_num)
        for y_num in range(1, nrows + 1):
            # page_position_list.append([x_num, y_num, ws.cell('%s%s' % (col, y_num)).value, sheet_name, excel_path])
            page_position_list.append([x_num, y_num, ws['%s%s' % (col, y_num)].value, sheet_name, excel_path])
    return page_position_list


def Main_function(mk_path):
    start = time.time()
    sheet_list = fun_sheet_list(mk_path)
    if type(sheet_list) is str:
        print(sheet_list)
    page_position_list = []
    print("正在计算文件夹下文件包含的最小单元格数量")
    for sheet_name in sheet_list:
        for excel_path in all_path(mk_path):
            page_position_list.extend(fun_page_position(excel_path, sheet_name))
    # 取得了整个大文件夹下所有数据的精确位置列表
    print("开始计算页数")
    for sheet_name in sheet_list:
        equal_sheet_list = []
        for page_position in page_position_list:
            if page_position[3] == sheet_name and page_position[2] == "第xxx页，共xxx页":
                equal_sheet_list.append(page_position)
        num = len(equal_sheet_list)
        page = 1
        for b in range(num):
            p = page_position_list.index(equal_sheet_list[b])
            if equal_sheet_list[b][2] == "第xxx页，共xxx页":
                equal_sheet_list[b][2] = "第" + str(page).zfill(3) + "页，共" + str(num).zfill(3) + "页"
                page_position_list[p] = equal_sheet_list[b]
                page = page + 1
        print("sheet {0} 共有 {1} 页".format(sheet_name, num))
    print("计算终了，开始进行页数信息写入")
    print("这里进度条显示的总数并不是相关页数")
    time.sleep(1)
    for index in tqdm(range(len(page_position_list))):
        page_position = page_position_list[index]
        if page_position[2] is not None and "页" in page_position[2]:
            wb = load_workbook(page_position[4])
            ws = wb[page_position[3]]
            ws.cell(row=int(page_position[1]), column=int(page_position[0]), value=page_position[2])
            wb.save(page_position[4])
    time.sleep(1)
    print("数据写入终了，请在关闭本脚本程序后人工检查。如有任何疑问，请于第一时间联系脚本编写人员。")
    c = time.time() - start
    print("脚本执行完成")
    print("脚本运行耗时:%0.2f S" % (c))
    input("点击回车退出程序")
    return "X"


if __name__ == '__main__':
    print("此脚本程序用于文件夹下Excel文件中页数的统计与修改，\n 要求Excel文件版本大于office2007")
    print("为避免不可预计的错误，请在运行前备份需要处理的文件夹")
    print("由于此脚本的检索逻辑是基于文件内‘第xxx页，共xxx页’，执行的。"
          "故请务必保证处理的文件是包含这个字段的，并不要将人工处理了一半的文件夹使用此程序处理")
    print("由于使用程序导致的文件损失，程序的编写者概不负责")
    print('为避免不可预计的错误，请在运行前备份需要处理的文件夹 !')
    print('为避免不可预计的错误，请在运行前备份需要处理的文件夹 !')
    print('为避免不可预计的错误，请在运行前备份需要处理的文件夹 !')
    print("重要的事情我说了四遍了，请务必备份")
    path = input("请输入需要整理的文件的文件夹路径，退出请输入‘X’：")
    while path != "X":
        print(path)
        if os.path.isdir(path):
            path = Main_function(path)
        else:
            path = input("您输入的不是一个有效的路径或命令，请重新输入：")
